import os
import time
from multiprocessing.pool import Pool

import grequests
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from utils.link import Link
from utils.sheet import ws_result, wb_result
from utils.utils import iterate_by_batch


def save_result_report(fileName, links, errors=None):
    ws_result.append([
        make_cell("Acceptor", bold=True),
        # make_cell("Anchor", bold=True),
        make_cell("Donor", bold=True),
        # make_cell("Has Anchor", bold=True),
        make_cell("Has Acceptor", bold=True),
        make_cell("Site Status", bold=True),
    ])

    for link in links:
        if link:
            ws_result.append(link.get_array_cells())

    if errors:
        ws_result.append([])
        ws_result.append([
            make_cell("Site", bold=True),
            make_cell("Error", bold=True),
        ])
        for error in errors:
            ws_result.append([error['site'], str(error['exception'])])

    backup_filename = 'sheets/{}'.format(time.strftime("Report-%Y-%m-%d %H-%M-%S.xlsx"))
    # wb_result.save(backup_filename)
    wb_result.save(fileName)


def make_cell(value, size=9, bold=False):
    cell = WriteOnlyCell(ws_result, value=value)
    cell.font = Font(name='Verdana', size=size, bold=bold)
    return cell


def exception_handler(request, exception):
    print('Exception', request, request.url, exception)


def check_acceptor_decorator(acceptor, anchor, donor):
    def check_acceptor(response, *args, **kwargs):
        print('GET', response.url, 'Status:', response.status_code, end=', ')

        # if response.status_code == 301:
        #     print('redirected to', response.headers.get('Location'), end=', kwargs:')
        # print(dir(response))

        link = Link(acceptor, anchor, donor)
        link.check(response.text, response.status_code)

        print('acceptor: ', acceptor in response.text, end=', ')
        print('anchor: ', anchor in response.text)
        response.link = link

        return response

    return check_acceptor


def check_url(rows, progressBar=None):
    print('Process', os.getpid(), 'Count of links', len(rows))
    chunk_size = 16
    chunk_count = 0
    chunk_results = []
    total_chunks_count = len(rows) / chunk_size

    chunks = iterate_by_batch(rows, chunk_size, None)

    print('Process', os.getpid(), 'Approximately Count of chunks', total_chunks_count)

    for chunk in chunks:
        chunk_count += 1
        print('Process', os.getpid(), 'Approximately Chunk: ', chunk_count, 'of', total_chunks_count)
        results = []

        for link in chunk:
            if not link:
                continue

            acceptor = link[0].value
            anchor = link[1].value
            donor = link[2].value

            if not acceptor or not donor or 'http' not in acceptor or 'http' not in donor: continue

            headers = {
                'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'}
            results.append(grequests.get(donor, headers=headers,
                                         hooks={'response': check_acceptor_decorator(acceptor, anchor, donor)}))

        results = grequests.map(results, exception_handler=exception_handler, size=16, gtimeout=12)

        if progressBar:
            progressBar.emit(chunk_count / total_chunks_count * 100)

        for result in results:
            if result != None:
                chunk_results.append(result.link)

    print('Finished: ', os.getpid())

    return chunk_results


def check_links_in_the_workbook(filename, progressBar=None, multi=True):
    print('CPU count:', os.cpu_count())
    wb = load_workbook(filename)
    ws = wb.active

    links = list(ws.iter_rows())

    batch_size = len(links) // os.cpu_count() + 1
    print('Total Links Count: {}, Batch Size: {}'.format(len(links), batch_size))

    if multi:
        batches = iterate_by_batch(links, batch_size, None)

        with Pool(processes=os.cpu_count()) as pool:
            pool_result = pool.map(check_url, batches, 1)

        result = []
        for r in pool_result:
            result += r
    else:
        result = check_url(links, progressBar)

    return result


if __name__ == '__main__':
    links = check_links_in_the_workbook('acceptor_links.xlsx')
    # links = check_links_in_the_workbook('examples/buy_cheap_essay (1).xlsx')
    # links = check_links_in_the_workbook('examples/college_papers_for_sale.xlsx')
    # links = check_links_in_the_workbook('examples/write_my_paper.xlsx')
    # links = check_links_in_the_workbook('examples/write_my_paper_for_me_cheap_essays_pay_for_papers (3).xlsx')
    save_result_report('acceptor_links_report.xlsx', links)
