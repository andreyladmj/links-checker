import urllib
from urllib.parse import urlencode, quote
from urllib.request import urlopen

import requests
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from utils.sheet import ws_result


class Link:
    def __init__(self, acceptor, anchor, donor):
        self.acceptor = acceptor.strip()
        self.anchor = anchor.strip()
        self.donor = donor
        self.has_anchor = False
        self.has_acceptor = False
        self.url_status = 0
        self.font = 'Verdana'
        self.author = 'Yanka=)'
        self.status_reason = 'OK'

    def check(self, text, status_code):
        text = text.lower()
        self.has_anchor = self.anchor.lower() in text
        self.has_acceptor = self.check_acceptor_in_text(text)
        self.url_status = status_code

        if not self.has_acceptor:
            self.try_to_find_a_problem(text)


    def check_acceptor_in_text(self, text):
        if self.acceptor.lower().replace('https://', '://') in text:
            return True

        if self.acceptor.lower().replace('http://', '://') in text:
            return True

        return self.acceptor.lower() in text

    def get_array_cells(self):
        acceptor_cell = WriteOnlyCell(ws_result, value=self.acceptor)
        anchor_cell = WriteOnlyCell(ws_result, value=self.anchor)
        donor_cell = WriteOnlyCell(ws_result, value=self.donor)

        if not self.has_anchor:
            acceptor_cell.fill = anchor_cell.fill = donor_cell.fill = PatternFill(start_color='FFfcf8e3',
                                                                                  end_color='FFfcf8e3',
                                                                                  fill_type='solid')

        if not self.has_acceptor:
            acceptor_cell.fill = anchor_cell.fill = donor_cell.fill = PatternFill(start_color='FFf2dede',
                                                                                  end_color='FFf2dede',
                                                                                  fill_type='solid')

        return [
            acceptor_cell,
            # anchor_cell,
            donor_cell,
            # self.get_anchor_cell(),
            self.get_donor_cell(),
            self.get_status_cell(),
        ]

    def get_anchor_cell(self):
        size = 8

        if self.has_anchor:
            cell = WriteOnlyCell(ws_result, value='OK')
            cell.fill = PatternFill(start_color='FFdff0d8', end_color='FFdff0d8', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='3c763d', bold=True)
            cell.comment = Comment(text="Good =)", author=self.author)
        else:
            cell = WriteOnlyCell(ws_result, value='NO')
            cell.fill = PatternFill(start_color='FFfcf8e3', end_color='FFfcf8e3', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='8a6d3b', bold=True)
            cell.comment = Comment(text="There are no anchor ({}) on the site: {}".format(self.anchor, self.acceptor),
                                   author=self.author)

        return cell

    def get_donor_cell(self):
        size = 9

        if self.has_acceptor:
            cell = WriteOnlyCell(ws_result, value='OK')
            cell.fill = PatternFill(start_color='FFdff0d8', end_color='FFdff0d8', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='3c763d', bold=True)
            cell.comment = Comment(text="Good =)", author=self.author)
        else:
            cell = WriteOnlyCell(ws_result, value='NO')
            cell.fill = PatternFill(start_color='FFf2dede', end_color='FFf2dede', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='a94442', bold=True)
            cell.comment = Comment(text="There are no acceptor ({}) on the site: {}".format(self.acceptor, self.donor),
                                   author=self.author)

        return cell

    def get_status_cell(self):
        size = 8

        if self.url_status < 400:
            cell = WriteOnlyCell(ws_result, value=self.status_reason)

            color = '3c763d' if self.has_acceptor else 'a94442'

            cell.font = Font(name=self.font, size=size, color=color, bold=True)
            cell.comment = Comment(text="Good =)", author=self.author)
        else:
            cell = WriteOnlyCell(ws_result, value="Site Access Problem")
            cell.font = Font(name=self.font, size=size, color='a94442', bold=True)
            cell.comment = Comment(text="Seems we have some troubles with the site: {}".format(self.acceptor),
                                   author=self.author)

        return cell

    def try_to_find_a_problem(self, text):
        acceptor = self.acceptor.lower()

        if acceptor.replace('http://', 'https://') in text:
            self.status_reason = 'Used Wrong Protocol (https)'

        if acceptor.replace('https://', 'http://') in text:
            self.status_reason = 'Used Wrong Protocol (http)'

        if quote(acceptor).lower() in text or quote(acceptor).replace('/', '%2F').lower() in text:
            self.has_acceptor = True


# if __name__ == '__main__':
#     acceptor = 'http://techiejerry.com/top-technology-trends-in-education/'
#     anchor = 'Top Technology Trends In Education'
#     donor = 'http://askheatherjarvis.com/forums/viewthread/10095/'
#
#     def g_response(response, *args, **kwargs):
#         print('g_response')
#
#         l = Link(acceptor, anchor, donor)
#         l.check(response.text, response.status_code)
#
#         print('has_anchor', l.has_anchor)
#         print('has_acceptor', l.has_acceptor)
#         print('problem', l.problem)
#
#     get_simple_page(donor, g_response)
