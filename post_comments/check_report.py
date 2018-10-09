import grequests
from openpyxl import load_workbook

headers = {
    'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'}

# from gevent import monkey
# monkey.patch_all()

success = []
fail = []
total = 0
current = 0


def site_response_callback_decorator(donor, acceptor, email, author, comment):
    def site_response_callback(response, **kwargs):
        global current, success, fail

        current += 1
        text = response.text.lower()
        print("{} of {}".format(current, total), donor, acceptor in text, email in text, author in text, comment in text)

        if comment in text:
            success.append([donor, acceptor, email, author, comment])
        else:
            fail.append([donor, acceptor, email, author, comment])

        return response

    return site_response_callback


def exception_handler(request, exception):
    global current
    current += 1
    print('Exception', request.url, exception)


if __name__ == '__main__':
    wb = load_workbook('report_last.xlsx')
    ws = wb.active
    links = map(lambda x: [i.value for i in x], ws.iter_rows())
    grequest_stack = []

    for link in links:
        donor = link[0]
        acceptor = link[1].lower()
        email = link[2].lower()
        author = link[3].lower()
        comment = link[4].lower()

        if 'http' not in donor: continue

        callback = site_response_callback_decorator(donor, acceptor, email, author, comment)
        grequest_stack.append(grequests.get(donor, headers=headers,
                                            hooks={'response': callback},
                                            timeout=10))

    total = len(grequest_stack)
    print('len grequest stack', len(grequest_stack))

    results = grequests.map(grequest_stack, exception_handler=exception_handler, size=16)
    print("\nSuccess: {}".format(len(success)))
    print(success)
