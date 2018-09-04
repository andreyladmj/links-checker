import requests
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from sheet import ws_result


class Link:
    def __init__(self, acceptor, anchor, donor):
        self.acceptor = acceptor
        self.anchor = anchor
        self.donor = donor
        self.has_anchor = False
        self.has_donor = False
        self.url_status = 0
        self.font = 'Verdana'

    def check(self):
        page = requests.get(self.donor)
        self.has_anchor = self.anchor in page.text
        self.has_donor = self.acceptor in page.text
        self.url_status = page.status_code
        print('Check', self.acceptor, page.status_code)

    def get_array_cells(self):
        acceptor_cell = WriteOnlyCell(ws_result, value=self.acceptor)
        anchor_cell = WriteOnlyCell(ws_result, value=self.anchor)
        donor_cell = WriteOnlyCell(ws_result, value=self.donor)

        if not self.has_anchor:
            acceptor_cell.fill = anchor_cell.fill = donor_cell.fill = PatternFill(start_color='FFfcf8e3',
                                                                                  end_color='FFfcf8e3',
                                                                                  fill_type='solid')

        if not self.has_donor:
            acceptor_cell.fill = anchor_cell.fill = donor_cell.fill = PatternFill(start_color='FFf2dede',
                                                                                  end_color='FFf2dede',
                                                                                  fill_type='solid')

        return [
            acceptor_cell,
            anchor_cell,
            donor_cell,
            self.get_anchor_cell(),
            self.get_donor_cell(),
            self.get_status_cell(),
        ]

    def get_anchor_cell(self):
        size = 8

        if self.has_anchor:
            cell = WriteOnlyCell(ws_result, value='OK')
            cell.fill = PatternFill(start_color='FFdff0d8', end_color='FFdff0d8', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='3c763d', bold=True)
            cell.comment = Comment(text="Good =)", author='Yana')
        else:
            cell = WriteOnlyCell(ws_result, value='NO')
            cell.fill = PatternFill(start_color='FFfcf8e3', end_color='FFfcf8e3', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='8a6d3b', bold=True)
            cell.comment = Comment(text="There are no anchor ({}) on the site: {}".format(self.anchor, self.acceptor),
                                   author='Yana')

        return cell

    def get_donor_cell(self):
        size = 9

        if self.has_donor:
            cell = WriteOnlyCell(ws_result, value='OK')
            cell.fill = PatternFill(start_color='FFdff0d8', end_color='FFdff0d8', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='3c763d', bold=True)
            cell.comment = Comment(text="Good =)", author='Yana')
        else:
            cell = WriteOnlyCell(ws_result, value='NO')
            cell.fill = PatternFill(start_color='FFf2dede', end_color='FFf2dede', fill_type='solid')
            cell.font = Font(name=self.font, size=size, color='a94442', bold=True)
            cell.comment = Comment(text="There are no acceptor ({}) on the site: {}".format(self.acceptor, self.donor),
                                   author='Yana')

        return cell

    def get_status_cell(self):
        size = 8

        if self.url_status == 200:
            cell = WriteOnlyCell(ws_result, value='OK')
            cell.font = Font(name=self.font, size=size, color='3c763d', bold=True)
            cell.comment = Comment(text="Good =)", author='Yana')
        else:
            cell = WriteOnlyCell(ws_result, value='Some problems')
            cell.font = Font(name=self.font, size=size, color='a94442', bold=True)
            cell.comment = Comment(text="Seems we have some troubles with the site: {}".format(self.acceptor),
                                   author='Yana')

        return cell
