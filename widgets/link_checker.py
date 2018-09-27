from multiprocessing import Queue
from os import cpu_count
from time import time

from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QProgressBar, QLabel, QHBoxLayout, QFileDialog, \
    QVBoxLayout, QGroupBox, QGridLayout, QScrollBar, QTextEdit
from openpyxl import load_workbook

from check_links import save_result_report
from utils.parse_xlsx import ParseXLSX
from utils.utils import iterate_by_batch
from widgets.lcwidget import LCWidget


class CheckAcceptors(LCWidget):

    def __init__(self, parent, number_of_threads):
        super().__init__(parent, number_of_threads)
        self.title = 'Acceptor Checker'
        self.all_processed_links = []
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        self.init_timer()

        self.horizontalGroupBox = QGroupBox("Checker")

        self.select_xlsx = QPushButton('Select XLSX')
        self.select_xlsx.clicked.connect(self.select_xlsx_dialog)

        self.export_xlsx_button = QPushButton('Export XLSX Report')
        self.export_xlsx_button.clicked.connect(self.export_xlsx)

        actions_layout = QGridLayout()
        actions_layout.setColumnStretch(1, 3)
        actions_layout.addWidget(self.select_xlsx,0,1)
        # actions_layout.addWidget(QPushButton('Export Logs'),0,1)
        actions_layout.addWidget(self.export_xlsx_button,0,2)

        vbox_layuot = self.getProcessesUI(ParseXLSX)

        actions_layout.addWidget(self.get_time_execution(), 1, 0, 1, 3)
        actions_layout.addLayout(vbox_layuot, 2, 0, 1, 3)

        self.horizontalGroupBox.setLayout(actions_layout)

        windowLayout = QVBoxLayout()
        windowLayout.addWidget(self.horizontalGroupBox)
        windowLayout.addWidget(self.qlogs)
        self.setLayout(windowLayout)

        self.show()

    def select_xlsx_dialog(self):
        file = self.openFileNameDialog()
        if file:
            self.qlogs.log("Selected file {}".format(file))
            self.parse_xlsx_file(file)

    def parse_xlsx_file(self, file):
        self.start_time = time()
        wb = load_workbook(file)
        ws = wb.active

        links = list(ws.iter_rows())

        batch_size = len(links) // self.processes + 1
        self.qlogs.log('Total Links Count: {}, Batch Size: {}'.format(len(links), batch_size))

        batches = iterate_by_batch(links, batch_size, None)

        for process, batch in zip(self.processes_list, batches):
            process.set_links(batch)
            process.start()

    def show_finished_program_info(self):
        all_links = []
        exceptions = []

        for process in self.processes_list:
            for link in process.processed_links:
                all_links.append(link)

            for exc in process.exceptions:
                exceptions.append(exc)

        self.qlogs.log('Processed: {} links'.format(len(all_links)))
        self.qlogs.log("")
        self.qlogs.log("Exceptions:")

        for ex in exceptions:
            self.qlogs.log("Site: {}, Exception: {}".format(ex['site'], ex['exception']))

        self.all_processed_links = all_links

    def export_xlsx(self):
        if not len(self.all_processed_links):
            return self.qlogs.log("There are no links to export")

        file = self.saveFileDialog()
        if file:
            self.qlogs.log("Saving to {} file".format(file))
            save_result_report(file, self.all_processed_links)
            self.qlogs.log("Saved!")