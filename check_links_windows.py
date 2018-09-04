import os
import time
from itertools import zip_longest
# from multiprocessing.pool import Pool
from multiprocessing.dummy import Pool

from PyQt5.QtCore import pyqtSignal, pyqtSlot
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from link import Link
from loader import ProgressBar_Dialog
from sheet import ws_result, wb_result


def iterate_by_batch(array_list, amount, fillvalue=None):
    args = [iter(array_list)] * amount
    return zip_longest(*args, fillvalue=fillvalue)


def parse_rows(rows):
    for row in rows:
        parse_row(row)

def parse_row(column):
    print('column', column)
    if not column:
        return None

    acceptor = column[0].value
    anchor = column[1].value
    donor = column[2].value

    if 'http' not in acceptor: return False

    link = Link(acceptor, anchor, donor)
    link.check()

    return link


def save_result_report(fileName, links):
    ws_result.append([
        make_cell("Acceptor", bold=True),
        make_cell("Anchor", bold=True),
        make_cell("Donor", bold=True),
        make_cell("Has Anchor", bold=True),
        make_cell("Has Donor", bold=True),
        make_cell("Site Status", bold=True),
    ])

    for link in links:
        if link:
            ws_result.append(link.get_array_cells())

    backup_filename = 'sheets/{}'.format(time.strftime("Report-%Y-%m-%d %H-%M-%S.xlsx"))
    wb_result.save(backup_filename)
    wb_result.save(fileName)


def make_cell(value, size=9, bold=False):
    cell = WriteOnlyCell(ws_result, value=value)
    cell.font = Font(name='Verdana', size=size, bold=bold)
    return cell


def check(filename='', progressBar=None):
    threads = 8
    all_links = []

    wb = load_workbook(filename)
    ws = wb.active
    count = 0


    rows = list(ws.iter_rows())
    batch_size = len(rows) // os.cpu_count() + 1

    batches = iterate_by_batch(rows, batch_size, None)


    with Pool(processes=os.cpu_count()) as pool:
        #for batch_rows in iterate_by_batch(ws.iter_rows(), threads, None):
        links = pool.map(parse_rows, batches, 1)
        print('links', links)
        all_links += links
        count += len(links)
        print(count)
        progress = count / ws.max_row * 100
        progressBar.emit(progress)

    return all_links


if __name__ == '__main__':
    check('college_papers_for_sale.xlsx')